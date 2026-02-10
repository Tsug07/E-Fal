import customtkinter as ctk
from tkinter import filedialog, messagebox
import tkinter as tk
import os
import threading
import time
import openpyxl
import pyautogui
from PIL import Image, ImageTk
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import psutil
from datetime import datetime

# Configuração do tema
ctk.set_appearance_mode("dark")  # Modes: "System" (standard), "Dark", "Light"
ctk.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

class TJRJInterface:
    def __init__(self, root):
        self.root = root
        self.root.title("E-Fal - Emissão de Certidão Falência")
        self.root.geometry("600x800")
        
        # Variáveis
        self.caminho_excel = tk.StringVar()
        self.pasta_destino = tk.StringVar(value=r"C:\Users\VM001\Documents\CNDs\FALENCIA")
        self.url = tk.StringVar(value="https://www3.tjrj.jus.br/CJE/certidao/judicial/visualizar?modelo=visualizar")
        self.executando = False
        self.driver = None
        self.log_file_path = None
        
        # Configurações globais
        self.MODELOS = {
            "E-falencia": {
                "colunas": ["Codigo", "Cliente", "CND", "Pedido"],
                "mensagem_padrao": "ONEmessage"
            }
        }
        
        self.setup_ui()
        
    def setup_ui(self):
        # Frame principal
        main_frame = ctk.CTkFrame(self.root)
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Título
        title_label = ctk.CTkLabel(main_frame, text="E-Fal", 
                                  font=ctk.CTkFont(size=24, weight="bold"))
        title_label.pack(pady=(20, 30))
        
        # Frame de configurações
        config_frame = ctk.CTkFrame(main_frame)
        config_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        # Label de configurações
        ctk.CTkLabel(config_frame, text="Configurações", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(15, 10))
        
        # Arquivo Excel
        excel_frame = ctk.CTkFrame(config_frame)
        excel_frame.pack(fill="x", padx=15, pady=5)
        
        ctk.CTkLabel(excel_frame, text="Arquivo Excel:").pack(anchor="w", padx=10, pady=(10, 5))
        
        excel_input_frame = ctk.CTkFrame(excel_frame)
        excel_input_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        self.excel_entry = ctk.CTkEntry(excel_input_frame, textvariable=self.caminho_excel, 
                                       placeholder_text="Selecione o arquivo Excel...")
        self.excel_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        ctk.CTkButton(excel_input_frame, text="Procurar", 
                     command=self.selecionar_excel, width=80).pack(side="right")
        
        # Pasta destino
        pasta_frame = ctk.CTkFrame(config_frame)
        pasta_frame.pack(fill="x", padx=15, pady=5)
        
        ctk.CTkLabel(pasta_frame, text="Pasta Destino:").pack(anchor="w", padx=10, pady=(10, 5))
        
        pasta_input_frame = ctk.CTkFrame(pasta_frame)
        pasta_input_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        self.pasta_entry = ctk.CTkEntry(pasta_input_frame, textvariable=self.pasta_destino)
        self.pasta_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        ctk.CTkButton(pasta_input_frame, text="Procurar", 
                     command=self.selecionar_pasta, width=80).pack(side="right")
        
        # URL
        url_frame = ctk.CTkFrame(config_frame)
        url_frame.pack(fill="x", padx=15, pady=(5, 15))
        
        ctk.CTkLabel(url_frame, text="URL TJRJ:").pack(anchor="w", padx=10, pady=(10, 5))
        
        self.url_entry = ctk.CTkEntry(url_frame, textvariable=self.url)
        self.url_entry.pack(fill="x", padx=10, pady=(0, 10))
        
        # Frame de controle
        control_frame = ctk.CTkFrame(main_frame)
        control_frame.pack(fill="x", padx=20, pady=(0, 20))
        
        # Botões
        buttons_frame = ctk.CTkFrame(control_frame)
        buttons_frame.pack(pady=20)
        
        self.btn_validar = ctk.CTkButton(buttons_frame, text="Validar Excel", 
                                        command=self.validar_excel, width=120)
        self.btn_validar.pack(side="left", padx=10)
        
        self.btn_iniciar = ctk.CTkButton(buttons_frame, text="Iniciar Processamento", 
                                        command=self.iniciar_processamento, width=150)
        self.btn_iniciar.pack(side="left", padx=10)
        
        self.btn_parar = ctk.CTkButton(buttons_frame, text="Parar", 
                                      command=self.parar_processamento, 
                                      state='disabled', width=80,
                                      fg_color="red", hover_color="darkred")
        self.btn_parar.pack(side="left", padx=10)
        
        # Progress bar
        self.progress = ctk.CTkProgressBar(main_frame, mode='indeterminate')
        self.progress.pack(fill="x", padx=20, pady=(0, 10))
        
        # Status
        self.status_label = ctk.CTkLabel(main_frame, text="Pronto", 
                                        text_color="green",
                                        font=ctk.CTkFont(size=12, weight="bold"))
        self.status_label.pack(pady=5)
        
        # Frame de dados
        dados_frame = ctk.CTkFrame(main_frame)
        dados_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        ctk.CTkLabel(dados_frame, text="Dados do Excel", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(15, 10))
        
        # Frame para o Treeview (usando tkinter nativo dentro do customtkinter)
        tree_frame = ctk.CTkFrame(dados_frame)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        # Treeview (usando tkinter nativo)
        tree_container = tk.Frame(tree_frame)
        tree_container.pack(fill="both", expand=True, padx=2, pady=2)
        
        # Configurar cores do treeview para combinar com o tema
        style = tk.ttk.Style()
        if ctk.get_appearance_mode() == "Dark":
            style.configure("Treeview", background="#2b2b2b", foreground="white", fieldbackground="#2b2b2b")
            style.configure("Treeview.Heading", background="#1f538d", foreground="white")
        
        self.tree = tk.ttk.Treeview(tree_container, columns=('Codigo', 'Cliente', 'Pedido'), 
                                   show='headings', height=8)
        self.tree.heading('Codigo', text='Código')
        self.tree.heading('Cliente', text='Cliente')
        self.tree.heading('Pedido', text='Pedido')
        
        self.tree.column('Codigo', width=100)
        self.tree.column('Cliente', width=300)
        self.tree.column('Pedido', width=150)
        
        # Scrollbar para treeview
        tree_scroll = tk.ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        tree_scroll.pack(side="right", fill="y")
        
        # Frame de log
        log_frame = ctk.CTkFrame(main_frame)
        log_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        ctk.CTkLabel(log_frame, text="Log de Execução", 
                    font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(15, 10))
        
        # Text widget para log
        self.log_text = ctk.CTkTextbox(log_frame, height=200)
        self.log_text.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
    def selecionar_excel(self):
        arquivo = filedialog.askopenfilename(
            title="Selecionar arquivo Excel",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if arquivo:
            self.caminho_excel.set(arquivo)
            
    def selecionar_pasta(self):
        pasta = filedialog.askdirectory(title="Selecionar pasta de destino")
        if pasta:
            self.pasta_destino.set(pasta)
            
    def atualizar_log(self, mensagem, cor=None):
        """Adiciona mensagem ao log com timestamp"""
        timestamp = datetime.now().strftime("[%H:%M:%S] ")
        mensagem_completa = f"{timestamp}{mensagem}\n"
        
        self.log_text.insert("end", mensagem_completa)
        self.log_text.see("end")
        
        # Atualiza o arquivo de log se existir
        if self.log_file_path and os.path.exists(self.log_file_path):
            with open(self.log_file_path, 'a', encoding='utf-8') as f:
                f.write(mensagem_completa)
        
        self.root.update_idletasks()
        
    def inicializar_arquivo_log(self):
        """Inicializa o arquivo de log com timestamp"""
        log_dir = os.path.join(os.path.dirname(__file__), 'TJRJ_Logs')
        os.makedirs(log_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file_path = os.path.join(log_dir, f"tjrj_log_{timestamp}.txt")
        with open(log_file_path, 'w', encoding='utf-8') as f:
            f.write(f"=== Log TJRJ Certidão de Falência - {timestamp} ===\n\n")
        return log_file_path
        
    def validar_excel(self):
        """Valida o arquivo Excel e carrega os dados na interface"""
        if not self.caminho_excel.get():
            messagebox.showerror("Erro", "Selecione um arquivo Excel primeiro!")
            return
            
        try:
            wb = openpyxl.load_workbook(self.caminho_excel.get())
            sheet = wb.active
            colunas_excel = [cell.value for cell in sheet[1]]
            colunas_esperadas = self.MODELOS["E-falencia"]["colunas"]
            
            if colunas_excel != colunas_esperadas:
                messagebox.showerror("Erro", f"O Excel não corresponde ao modelo E-falencia.\nEsperado: {colunas_esperadas}\nEncontrado: {colunas_excel}")
                return False
                
            # Limpa a treeview
            for item in self.tree.get_children():
                self.tree.delete(item)
                
            # Carrega os dados
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 4:
                    codigo, cliente, cnd, pedido = row[:4]
                    if codigo and pedido:  # Verifica se os campos obrigatórios estão preenchidos
                        self.tree.insert('', 'end', values=(codigo, cliente, pedido))
                        
            messagebox.showinfo("Sucesso", "Excel validado com sucesso!")
            self.atualizar_log("Excel validado com sucesso!")
            return True
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao validar Excel: {str(e)}")
            self.atualizar_log(f"Erro ao validar Excel: {str(e)}")
            return False
            
    def ler_dados_excel(self, caminho_excel, modelo, linha_inicial=2):
        """Lê os dados do Excel"""
        try:
            wb = openpyxl.load_workbook(caminho_excel)
            sheet = wb.active
            dados = {}
            colunas = self.MODELOS[modelo]["colunas"]
            
            for row in sheet.iter_rows(min_row=linha_inicial, values_only=True):
                if row and len(row) >= len(colunas):
                    codigo, cliente, cnd, pedido = row[:4]
                    if codigo and pedido:
                        dados[codigo] = {
                            'codigo': codigo,
                            'cliente': cliente,
                            'pedido': pedido,
                        }
                        
            return dados if dados else None
            
        except Exception as e:
            self.atualizar_log(f"Erro ao ler Excel: {str(e)}")
            return None
            
    def encerrar_processos_chrome(self):
        """Encerra processos do Chrome de automação"""
        for proc in psutil.process_iter(['name', 'cmdline']):
            if proc.info['name'] == 'chrome.exe':
                try:
                    cmdline = proc.info['cmdline'] or []
                    cmdline_str = ' '.join(cmdline)
                    if '--user-data-dir=C:\\PerfisChrome\\automacao' in cmdline_str and '--profile-directory=Profile 1' in cmdline_str:
                        proc.terminate()
                        self.atualizar_log(f"Processo Chrome de automação (PID: {proc.pid}) encerrado.")
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
        time.sleep(2)
        
    def abrir_chrome_com_url(self, url):
        """Abre o Chrome com a URL especificada"""
        self.encerrar_processos_chrome()
        user_data_dir = r"C:\PerfisChrome\automacao"
        profile_dir = os.path.join(user_data_dir, "Profile 1")
        
        if not os.path.exists(profile_dir):
            self.atualizar_log(f"Perfil 'Profile 1' não encontrado em {user_data_dir}.")
            self.atualizar_log("Um novo perfil será criado. Por favor, faça login na página aberta para continuar.")
        
        chrome_options = Options()
        chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
        chrome_options.add_argument("--profile-directory=Profile 1")
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--disable-translate")
        chrome_options.add_argument("--lang=pt-BR")
        chrome_options.add_argument("--enable-javascript")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        try:
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
            driver.set_page_load_timeout(180)
            driver.get(url)
            self.atualizar_log(f"Chrome aberto com a URL: {url}")
            time.sleep(3)
            self.atualizar_log(f"Página acessada com sucesso: {driver.title}")
            return driver
        except Exception as e:
            self.atualizar_log(f"Erro ao abrir o Chrome: {str(e)}")
            return None
            
    def salvar_arquivo(self, info, data_vencimento):
        """Lida com a janela 'Salvar como'"""
        try:
            time.sleep(2)
            
            # Pressiona Tab 6 vezes para focar no campo de caminho/nome
            for _ in range(6):
                pyautogui.press('tab')
                time.sleep(0.2)
            
            pyautogui.press('enter')
            time.sleep(0.5)
            
            # Digita o caminho da pasta
            pyautogui.write(self.pasta_destino.get())
            self.atualizar_log(f"Caminho da pasta inserido: {self.pasta_destino.get()}")
            
            pyautogui.press('enter')
            time.sleep(3)
            
            # Pressiona Tab 6 vezes novamente para focar no campo de nome do arquivo
            for _ in range(6):
                pyautogui.press('tab')
                time.sleep(0.2)
            
            # Monta o nome do arquivo
            nome_arquivo = f"{info['codigo']} - CND FALENCIA - {data_vencimento}"
            pyautogui.write(nome_arquivo.upper())
            self.atualizar_log(f"Nome do arquivo inserido: {nome_arquivo}")
            
            pyautogui.press('enter')
            self.atualizar_log("Arquivo salvo com sucesso.")
            time.sleep(1)
            
            return True
        except Exception as e:
            self.atualizar_log(f"Erro ao salvar arquivo: {str(e)}")
            return False
            
    def formatar_data_vencimento(self, data):
        """Converte a data de vencimento de DD/MM/YYYY para DD.MM.YYYY"""
        try:
            return data.replace('/', '.')
        except Exception as e:
            self.atualizar_log(f"Erro ao formatar data de vencimento: {str(e)}")
            return data
            
    def processar_pedido(self, driver, info):
        """Processa um pedido individual"""
        try:
            # Preenche o campo de requerimento
            campo_requerimento = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="NumeroCertidao"]'))
            )
            campo_requerimento.clear()
            campo_requerimento.send_keys(info['pedido'])
            self.atualizar_log(f"Campo de requerimento preenchido com: {info['pedido']}")
            
            # Clica no botão Visualizar
            botao_visualizar = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/div[3]/div/div[1]/div[1]/form/div[2]/div[2]/button'))
            )
            botao_visualizar.click()
            self.atualizar_log("Botão 'Visualizar' clicado com sucesso.")
            
            # Aguarda a tabela carregar e extrai a data de vencimento
            data_vencimento_elemento = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div[3]/div/div[1]/div[1]/table/tbody/tr[1]/td[5]'))
            )
            data_vencimento = data_vencimento_elemento.text.strip()
            data_vencimento_formatada = self.formatar_data_vencimento(data_vencimento)
            self.atualizar_log(f"Data de vencimento extraída: {data_vencimento} (Formatada: {data_vencimento_formatada})")
            
            # Clica no link para abrir a página de visualização do arquivo
            link_arquivo = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/div[3]/div/div[1]/div[1]/table/tbody/tr[1]/td[1]/a'))
            )
            link_arquivo.click()
            self.atualizar_log("Link do arquivo clicado com sucesso.")
            
            # Aguarda e aciona o download
            self.atualizar_log("Tentando acionar download com Ctrl + S.")
            time.sleep(4)
            pyautogui.hotkey('ctrl', 's')
            self.atualizar_log("Comando Ctrl + S executado com sucesso.")
            
            time.sleep(4)
            
            # Lida com a janela "Salvar como"
            if self.salvar_arquivo(info, data_vencimento_formatada):
                self.atualizar_log(f"Arquivo salvo para pedido {info['pedido']}.")
            else:
                self.atualizar_log(f"Falha ao salvar arquivo para pedido {info['pedido']}.")
                return None
                
            # Fecha a aba de visualização
            pyautogui.hotkey('ctrl', 'w')
            self.atualizar_log("Aba de visualização fechada com sucesso.")
            time.sleep(1)
            
            return data_vencimento
            
        except Exception as e:
            self.atualizar_log(f"Erro ao processar pedido {info['pedido']}: {str(e)}")
            return None
            
    def processar_dados(self):
        """Função principal de processamento"""
        try:
            # Inicializa o log
            self.log_file_path = self.inicializar_arquivo_log()
            
            # Valida configurações
            if not self.caminho_excel.get():
                messagebox.showerror("Erro", "Selecione um arquivo Excel!")
                return
                
            if not os.path.exists(self.pasta_destino.get()):
                messagebox.showerror("Erro", "Pasta de destino não existe!")
                return
                
            # Lê os dados do Excel
            dados = self.ler_dados_excel(self.caminho_excel.get(), "E-falencia")
            if not dados:
                self.atualizar_log("Nenhum dado para processar no Excel.")
                return
                
            # Abre o Chrome
            self.atualizar_log("Abrindo Chrome...")
            self.driver = self.abrir_chrome_com_url(self.url.get())
            if not self.driver:
                self.atualizar_log("Não foi possível abrir o Chrome.")
                return
                
            self.atualizar_log("Iniciando processamento dos pedidos...")
            
            # Processa cada pedido
            total_pedidos = len(dados)
            for i, (codigo, info) in enumerate(dados.items(), 1):
                if not self.executando:
                    break
                    
                self.atualizar_log(f"Processando {i}/{total_pedidos} - Código: {codigo}, Cliente: {info['cliente']}, Pedido: {info['pedido']}")
                
                # Processa o pedido
                data_vencimento = self.processar_pedido(self.driver, info)
                if data_vencimento:
                    self.atualizar_log(f"Processamento concluído para pedido {info['pedido']}. Data de vencimento: {data_vencimento}")
                else:
                    self.atualizar_log(f"Falha no processamento do pedido {info['pedido']}.")
                
                # Volta para a página inicial
                if self.executando:
                    self.driver.get(self.url.get())
                    time.sleep(3)
                    
            self.atualizar_log("Processamento finalizado!")
            
        except Exception as e:
            self.atualizar_log(f"Erro durante o processamento: {str(e)}")
            messagebox.showerror("Erro", f"Erro durante o processamento: {str(e)}")
            
        finally:
            # Limpa recursos
            if self.driver:
                self.driver.quit()
                self.driver = None
                
            self.executando = False
            self.progress.stop()
            self.btn_iniciar.configure(state='normal')
            self.btn_parar.configure(state='disabled')
            self.status_label.configure(text="Processamento finalizado", text_color="green")
            
    def iniciar_processamento(self):
        """Inicia o processamento em uma thread separada"""
        if not self.caminho_excel.get():
            messagebox.showerror("Erro", "Selecione um arquivo Excel primeiro!")
            return
            
        self.executando = True
        self.progress.start()
        self.btn_iniciar.configure(state='disabled')
        self.btn_parar.configure(state='normal')
        self.status_label.configure(text="Processando...", text_color="orange")
        
        # Inicia o processamento em uma thread separada
        thread = threading.Thread(target=self.processar_dados)
        thread.daemon = True
        thread.start()
        
    def parar_processamento(self):
        """Para o processamento"""
        self.executando = False
        self.progress.stop()
        self.btn_iniciar.configure(state='normal')
        self.btn_parar.configure(state='disabled')
        self.status_label.configure(text="Processamento interrompido", text_color="red")
        self.atualizar_log("Processamento interrompido pelo usuário.")
        
        if self.driver:
            self.driver.quit()
            self.driver = None

def main():
    root = ctk.CTk()
    app = TJRJInterface(root)
    root.mainloop()

if __name__ == "__main__":
    main()