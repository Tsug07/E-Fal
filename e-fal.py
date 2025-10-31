import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import os
import threading
import time
import openpyxl
import pyautogui
from PIL import Image, ImageTk
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import psutil
from datetime import datetime
import sys

class TJRJInterface:
    def __init__(self, root):
        self.root = root
        self.root.title("E-Fal - Emissão de Certidão Falência")
        self.root.geometry("500x700")
        self.root.configure(bg='#1e1e1e')  # Fundo escuro
        
        # Configura o ícone da janela
        if getattr(sys, 'frozen', False):
            # Se for um executável, usa o diretório temporário do PyInstaller
            icon_path = os.path.join(sys._MEIPASS, "assets", "E-Fal.ico")
        else:
            # Se for no ambiente de desenvolvimento, usa o caminho relativo
            icon_path = os.path.join(os.path.dirname(__file__), "assets", "E-Fal.ico")
        try:
            self.root.iconbitmap(icon_path)
        except tk.TclError as e:
            print(f"Erro ao carregar ícone: {e}")
            
        # Variáveis
        self.caminho_excel = tk.StringVar()
        self.pasta_destino = tk.StringVar(value=r"C:\Users\VM001\Documents\CNDs\FALENCIA")
        self.url = tk.StringVar(value="https://www3.tjrj.jus.br/CJE/certidao/judicial/visualizar?modelo=visualizar")
        self.executando = False
        self.driver = None
        self.log_file_path = None
        
        # Configurações globais
        self.MODELOS = {
            "E-falencia": {
                "colunas": ["Codigo", "Cliente", "CND", "Pedido"],
                "mensagem_padrao": "ONEmessage"
            }
        }
        
        self.setup_ui()
        # Registra erro de ícone no log, se houver
        if 'icon_path' in locals() and not os.path.exists(icon_path):
            self.atualizar_log(f"Ícone não encontrado em: {icon_path}")
            
    def setup_ui(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configuração do grid
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        
        if getattr(sys, 'frozen', False):
            # Se for um executável, usa o diretório temporário do PyInstaller
            logo_path = os.path.join(sys._MEIPASS, "assets", "E-Fal.png")
        else:
            # Se for no ambiente de desenvolvimento, usa o caminho relativo
            logo_path = os.path.join(os.path.dirname(__file__), "assets", "E-Fal.png")
        logo_image = Image.open(logo_path)
        logo_image = logo_image.resize((120, 120), Image.LANCZOS)  # ajuste o tamanho conforme necessário
        self.logo_photo = ImageTk.PhotoImage(logo_image)

       # Frame horizontal para logo e título
        header_frame = tk.Frame(main_frame, bg="#1e1e1e")
        header_frame.grid(row=0, column=0, columnspan=5, pady=(0, 10), sticky="w")
        # Logo
        logo_label = tk.Label(header_frame, image=self.logo_photo, bg="#1e1e1e")
        logo_label.pack(side="left", padx=(0, 10))

        # Título
        title_label = tk.Label(header_frame, text="                  E-Fal                                  ",
                            font=('Arial', 16, 'bold'), fg="white", bg="#1e1e1e")
        title_label.pack(side="left")
        
        # Frame de configurações
        config_frame = ttk.LabelFrame(main_frame, text="Configurações", padding="10")
        config_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        config_frame.columnconfigure(1, weight=1)
        
        # Arquivo Excel
        ttk.Label(config_frame, text="Arquivo Excel:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(config_frame, textvariable=self.caminho_excel, width=60).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(5, 5))
        ttk.Button(config_frame, text="Procurar", command=self.selecionar_excel).grid(row=0, column=2, padx=(5, 0))
        
        # Pasta destino
        ttk.Label(config_frame, text="Pasta Destino:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(config_frame, textvariable=self.pasta_destino, width=60).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(5, 5))
        ttk.Button(config_frame, text="Procurar", command=self.selecionar_pasta).grid(row=1, column=2, padx=(5, 0))
        
        # URL
        ttk.Label(config_frame, text="URL TJRJ:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(config_frame, textvariable=self.url, width=60).grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(5, 5))
        
        # Frame de controle
        control_frame = ttk.Frame(main_frame)
        control_frame.grid(row=2, column=0, columnspan=3, pady=10)
        
        # Botões
        self.btn_validar = ttk.Button(control_frame, text="Validar Excel", command=self.validar_excel)
        self.btn_validar.grid(row=0, column=0, padx=5)
        
        self.btn_iniciar = ttk.Button(control_frame, text="Iniciar Processamento", command=self.iniciar_processamento)
        self.btn_iniciar.grid(row=0, column=1, padx=5)
        
        self.btn_parar = ttk.Button(control_frame, text="Parar", command=self.parar_processamento, state='disabled')
        self.btn_parar.grid(row=0, column=2, padx=5)
        
        # Progress bar
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        # Status
        self.status_label = ttk.Label(main_frame, text="Pronto", foreground="green")
        self.status_label.grid(row=4, column=0, columnspan=3, pady=5)
        
        # Frame de dados
        dados_frame = ttk.LabelFrame(main_frame, text="Dados do Excel", padding="10")
        dados_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        dados_frame.columnconfigure(0, weight=1)
        dados_frame.rowconfigure(0, weight=1)
        
        # Treeview para dados
        self.tree = ttk.Treeview(dados_frame, columns=('Codigo', 'Cliente', 'Pedido'), show='headings', height=8)
        self.tree.heading('Codigo', text='Código')
        self.tree.heading('Cliente', text='Cliente')
        self.tree.heading('Pedido', text='Pedido')
        
        self.tree.column('Codigo', width=100)
        self.tree.column('Cliente', width=300)
        self.tree.column('Pedido', width=150)
        
        # Scrollbar para treeview
        tree_scroll = ttk.Scrollbar(dados_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        tree_scroll.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Frame de log
        log_frame = ttk.LabelFrame(main_frame, text="Log de Execução", padding="10")
        log_frame.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # Text widget para log
        self.log_text = scrolledtext.ScrolledText(log_frame, height=12, width=80)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configurar redimensionamento
        main_frame.rowconfigure(5, weight=1)
        main_frame.rowconfigure(6, weight=1)
        
    def selecionar_excel(self):
        arquivo = filedialog.askopenfilename(
            title="Selecionar arquivo Excel",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if arquivo:
            self.caminho_excel.set(arquivo)
            
    def selecionar_pasta(self):
        pasta = filedialog.askdirectory(title="Selecionar pasta de destino")
        if pasta:
            self.pasta_destino.set(pasta)
            
    def atualizar_log(self, mensagem, cor=None):
        """Adiciona mensagem ao log com timestamp"""
        timestamp = datetime.now().strftime("[%H:%M:%S] ")
        mensagem_completa = f"{timestamp}{mensagem}\n"
        
        self.log_text.insert(tk.END, mensagem_completa)
        self.log_text.see(tk.END)
        
        # Atualiza o arquivo de log se existir
        if self.log_file_path and os.path.exists(self.log_file_path):
            with open(self.log_file_path, 'a', encoding='utf-8') as f:
                f.write(mensagem_completa)
        
        self.root.update_idletasks()
        
    def inicializar_arquivo_log(self):
        """Inicializa o arquivo de log com timestamp"""
        log_dir = os.path.join(os.path.dirname(__file__), 'TJRJ_Logs')
        os.makedirs(log_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file_path = os.path.join(log_dir, f"tjrj_log_{timestamp}.txt")
        with open(log_file_path, 'w', encoding='utf-8') as f:
            f.write(f"=== Log TJRJ Certidão de Falência - {timestamp} ===\n\n")
        return log_file_path
        
    def validar_excel(self):
        """Valida o arquivo Excel e carrega os dados na interface"""
        if not self.caminho_excel.get():
            messagebox.showerror("Erro", "Selecione um arquivo Excel primeiro!")
            return
            
        try:
            wb = openpyxl.load_workbook(self.caminho_excel.get())
            sheet = wb.active
            colunas_excel = [cell.value for cell in sheet[1]]
            colunas_esperadas = self.MODELOS["E-falencia"]["colunas"]
            
            if colunas_excel != colunas_esperadas:
                messagebox.showerror("Erro", f"O Excel não corresponde ao modelo E-falencia.\nEsperado: {colunas_esperadas}\nEncontrado: {colunas_excel}")
                return False
                
            # Limpa a treeview
            for item in self.tree.get_children():
                self.tree.delete(item)
                
            # Carrega os dados
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 4:
                    codigo, cliente, cnd, pedido = row[:4]
                    if codigo and pedido:  # Verifica se os campos obrigatórios estão preenchidos
                        self.tree.insert('', 'end', values=(codigo, cliente, pedido))
                        
            messagebox.showinfo("Sucesso", "Excel validado com sucesso!")
            self.atualizar_log("Excel validado com sucesso!")
            return True
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao validar Excel: {str(e)}")
            self.atualizar_log(f"Erro ao validar Excel: {str(e)}")
            return False
            
    def ler_dados_excel(self, caminho_excel, modelo, linha_inicial=2):
        """Lê os dados do Excel"""
        try:
            wb = openpyxl.load_workbook(caminho_excel)
            sheet = wb.active
            dados = []
            colunas = self.MODELOS[modelo]["colunas"]
            
            for row in sheet.iter_rows(min_row=linha_inicial, values_only=True):
                if row and len(row) >= len(colunas):
                    codigo, cliente, cnd, pedido = row[:4]
                    if codigo and pedido:
                        dados.append({
                            'codigo': codigo,
                            'cliente': cliente,
                            'cnd': cnd,
                            'pedido': pedido,
                        })
                        
            return dados if dados else None
            
        except Exception as e:
            self.atualizar_log(f"Erro ao ler Excel: {str(e)}")
            return None
            
    def encerrar_processos_chrome(self):
        """Encerra processos do Chrome de automação"""
        for proc in psutil.process_iter(['name', 'cmdline']):
            if proc.info['name'] == 'chrome.exe':
                try:
                    cmdline = proc.info['cmdline'] or []
                    cmdline_str = ' '.join(cmdline)
                    if '--user-data-dir=C:\\PerfisChrome\\automacao' in cmdline_str and '--profile-directory=Profile 1' in cmdline_str:
                        proc.terminate()
                        self.atualizar_log(f"Processo Chrome de automação (PID: {proc.pid}) encerrado.")
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
        time.sleep(2)
        
    def abrir_chrome_com_url(self, url):
        """Abre o Chrome com a URL especificada"""
        self.encerrar_processos_chrome()
        user_data_dir = r"C:\PerfisChrome\automacao"
        profile_dir = os.path.join(user_data_dir, "Profile 1")
        
        if not os.path.exists(profile_dir):
            self.atualizar_log(f"Perfil 'Profile 1' não encontrado em {user_data_dir}.")
            self.atualizar_log("Um novo perfil será criado. Por favor, faça login na página aberta para continuar.")
        
        chrome_options = Options()
        chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
        chrome_options.add_argument("--profile-directory=Profile 1")
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--disable-translate")
        chrome_options.add_argument("--lang=pt-BR")
        chrome_options.add_argument("--enable-javascript")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        try:
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
            driver.set_page_load_timeout(180)
            driver.get(url)
            self.atualizar_log(f"Chrome aberto com a URL: {url}")
            time.sleep(3)
            self.atualizar_log(f"Página acessada com sucesso: {driver.title}")
            return driver
        except Exception as e:
            self.atualizar_log(f"Erro ao abrir o Chrome: {str(e)}")
            return None
            
    def salvar_arquivo(self, info, data_vencimento):
        """Lida com a janela 'Salvar como'"""
        try:
            time.sleep(2)
            
            # Pressiona Tab 6 vezes para focar no campo de caminho/nome
            for _ in range(6):
                pyautogui.press('tab')
                time.sleep(0.2)
            
            pyautogui.press('enter')
            time.sleep(0.5)
            
            # Digita o caminho da pasta
            pyautogui.write(self.pasta_destino.get())
            self.atualizar_log(f"Caminho da pasta inserido: {self.pasta_destino.get()}")
            
            pyautogui.press('enter')
            time.sleep(3)
            
            # Pressiona Tab 6 vezes novamente para focar no campo de nome do arquivo
            for _ in range(6):
                pyautogui.press('tab')
                time.sleep(0.2)
            
            # Monta o nome do arquivo
            nome_arquivo = f"{info['codigo']} - {info['cnd']} - {data_vencimento}"
            pyautogui.write(nome_arquivo)
            self.atualizar_log(f"Nome do arquivo inserido: {nome_arquivo}")
            
            pyautogui.press('enter')
            self.atualizar_log("Arquivo salvo com sucesso.")
            time.sleep(1)
            
            return True
        except Exception as e:
            self.atualizar_log(f"Erro ao salvar arquivo: {str(e)}")
            return False
            
    def formatar_data_vencimento(self, data):
        """Converte a data de vencimento de DD/MM/YYYY para DD.MM.YYYY"""
        try:
            return data.replace('/', '.')
        except Exception as e:
            self.atualizar_log(f"Erro ao formatar data de vencimento: {str(e)}")
            return data
            
    def processar_pedido(self, driver, info):
        """Processa um pedido individual"""
        try:
            # Preenche o campo de requerimento
            campo_requerimento = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="NumeroCertidao"]'))
            )
            campo_requerimento.clear()
            campo_requerimento.send_keys(info['pedido'])
            self.atualizar_log(f"Campo de requerimento preenchido com: {info['pedido']}")
            
            # Clica no botão Visualizar
            botao_visualizar = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/div[3]/div/div[1]/div[1]/form/div[2]/div[2]/button'))
            )
            botao_visualizar.click()
            self.atualizar_log("Botão 'Visualizar' clicado com sucesso.")
            
            # Aguarda a tabela carregar e extrai a data de vencimento
            data_vencimento_elemento = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div/div[3]/div/div[1]/div[1]/table/tbody/tr[1]/td[6]'))
            )
            data_vencimento = data_vencimento_elemento.text.strip()
            data_vencimento_formatada = self.formatar_data_vencimento(data_vencimento)
            self.atualizar_log(f"Data de vencimento extraída: {data_vencimento} (Formatada: {data_vencimento_formatada})")
            
            if data_vencimento_formatada == "":
                return data_vencimento_formatada
            # Clica no link para abrir a página de visualização do arquivo
            link_arquivo = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/div[3]/div/div[1]/div[1]/table/tbody/tr[1]/td[1]/a'))
            )
            link_arquivo.click()
            self.atualizar_log("Link do arquivo clicado com sucesso.")
            
            # Aguarda e aciona o download
            self.atualizar_log("Tentando acionar download com Ctrl + S.")
            time.sleep(4)
            pyautogui.hotkey('ctrl', 's')
            self.atualizar_log("Comando Ctrl + S executado com sucesso.")
            
            time.sleep(4)
            
            # Lida com a janela "Salvar como"
            if self.salvar_arquivo(info, data_vencimento_formatada):
                self.atualizar_log(f"Arquivo salvo para pedido {info['pedido']}.")
            else:
                self.atualizar_log(f"Falha ao salvar arquivo para pedido {info['pedido']}.")
                return None
                
            # Fecha a aba de visualização
            pyautogui.hotkey('ctrl', 'w')
            self.atualizar_log("Aba de visualização fechada com sucesso.")
            time.sleep(1)
            
            return data_vencimento
            
        except Exception as e:
            self.atualizar_log(f"Erro ao processar pedido {info['pedido']}: {str(e)}")
            return None
            
    def processar_dados(self):
        """Função principal de processamento"""
        try:
            # Inicializa o log
            self.log_file_path = self.inicializar_arquivo_log()
            
            # Valida configurações
            if not self.caminho_excel.get():
                messagebox.showerror("Erro", "Selecione um arquivo Excel!")
                return
                
            if not os.path.exists(self.pasta_destino.get()):
                messagebox.showerror("Erro", "Pasta de destino não existe!")
                return
                
            # Lê os dados do Excel
            dados = self.ler_dados_excel(self.caminho_excel.get(), "E-falencia")
            if not dados:
                self.atualizar_log("Nenhum dado para processar no Excel.")
                return
                
            # Abre o Chrome
            self.atualizar_log("Abrindo Chrome...")
            self.driver = self.abrir_chrome_com_url(self.url.get())
            if not self.driver:
                self.atualizar_log("Não foi possível abrir o Chrome.")
                return
                
            self.atualizar_log("Iniciando processamento dos pedidos...")
            
            # Processa cada pedido
            total_pedidos = len(dados)
            for i, info in enumerate(dados, 1):
                if not self.executando:
                    break
                    
                self.atualizar_log(f"Processando {i}/{total_pedidos} - Código: {info['codigo']}, Cliente: {info['cliente']}, Pedido: {info['pedido']}")
                
                # Processa o pedido
                data_vencimento = self.processar_pedido(self.driver, info)
                if data_vencimento:
                    self.atualizar_log(f"Processamento concluído para pedido {info['pedido']}. Data de vencimento: {data_vencimento}")
                else:
                    self.atualizar_log(f"Falha no processamento do pedido {info['pedido']}.")
                
                # Volta para a página inicial
                if self.executando:
                    self.driver.get(self.url.get())
                    time.sleep(3)
                    
            self.atualizar_log("Processamento finalizado!")
            
        except Exception as e:
            self.atualizar_log(f"Erro durante o processamento: {str(e)}")
            messagebox.showerror("Erro", f"Erro durante o processamento: {str(e)}")
            
        finally:
            # Limpa recursos
            if self.driver:
                self.driver.quit()
                self.driver = None
                
            self.executando = False
            self.progress.stop()
            self.btn_iniciar.config(state='normal')
            self.btn_parar.config(state='disabled')
            self.status_label.config(text="Processamento finalizado", foreground="green")
            
    def iniciar_processamento(self):
        """Inicia o processamento em uma thread separada"""
        if not self.caminho_excel.get():
            messagebox.showerror("Erro", "Selecione um arquivo Excel primeiro!")
            return
            
        self.executando = True
        self.progress.start()
        self.btn_iniciar.config(state='disabled')
        self.btn_parar.config(state='normal')
        self.status_label.config(text="Processando...", foreground="blue")
        
        # Inicia o processamento em uma thread separada
        thread = threading.Thread(target=self.processar_dados)
        thread.daemon = True
        thread.start()
        
    def parar_processamento(self):
        """Para o processamento"""
        self.executando = False
        self.progress.stop()
        self.btn_iniciar.config(state='normal')
        self.btn_parar.config(state='disabled')
        self.status_label.config(text="Processamento interrompido", foreground="red")
        self.atualizar_log("Processamento interrompido pelo usuário.")
        
        if self.driver:
            self.driver.quit()
            self.driver = None

def main():
    root = tk.Tk()
    app = TJRJInterface(root)
    root.mainloop()

if __name__ == "__main__":
    main()